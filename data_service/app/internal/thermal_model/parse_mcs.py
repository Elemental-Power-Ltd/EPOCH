"""
Functions to parse a specific set of MCS room survey spreadsheets.

These are likely to be fragile, so watch out for the format changing over time.
"""

import datetime
import itertools
import pathlib
from collections import defaultdict
from collections.abc import MutableMapping
from typing import TypedDict

import openpyxl
import openpyxl.worksheet
import openpyxl.worksheet.worksheet

from ..utils.conversions import try_convert_float


class MCSRoomMetadata(TypedDict):
    """Room metadata from an MCS survey."""

    room_name: str | None
    ach: float | None
    level: str | None


def extract_design_details(file_path: pathlib.Path) -> dict[str, dict[str, dict[str, str | float]]]:
    """
    Extract design details of a specific building.

    :param file_path: Path to the Excel file
    :return: Nested dictionary of U-Values and Construction materials by storey
    """
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True, keep_links=False)

    sheet = wb["Design Details"]

    result: MutableMapping[str, dict[str, dict[str, str | float]]] = defaultdict(dict)
    read_mode = False
    storey = None

    # These max rows and cols aren't perfect, they just rule out a chunk of the file
    # that we know we're not interested in and the parser handles the rest.
    for row in sheet.iter_rows(values_only=True, max_col=12, min_row=30):
        if row[1] == "Property assumed U Values":
            # Start reading for any row after this one as we're
            # in the fabric section.
            read_mode = True
        elif row[1] == "Space Heating":
            # But stop reading here, as we're out of the interesting section.
            # Note that "Space Heating" shows up twice (once in the aggregrates at the top, once at the bottom)
            # but they're both out of the `read_mode` block so we're fine.
            read_mode = False

        # Skip rows that aren't in the block we care about, or blank rows.
        if (not read_mode) or (not any(row)):
            continue

        if any("Construction" in str(cell) for cell in row):
            # We've found a header row for a particular sub-type of the building.
            # Extract the storey, which we'll use until we find another one,
            # and stop processing this row.
            storey = str(row[1]).strip() if row[1] is not None else None
            continue

        element = str(row[1])
        u_value = try_convert_float(row[2]) if row[2] else None
        construction = str(row[3]) if row[3] else None
        if storey is not None and u_value is not None and construction is not None:
            result[storey][element] = {"U-Value": u_value, "Construction": construction}

    return dict(result)


def parse_single_room(sheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, list[dict[str, float]]]:
    """
    Parse a single MCS worksheet corresponding to room dimensions.

    Creates a list of different elements of a give type, e.g. there may be four external wall elements.

    Parameters
    ----------
    sheet
        Openpyxl worksheet from an MCS survey

    Returns
    -------
        Dictionary of {element_type: list[element sizes]}
    """

    def parse_row(row: tuple[str | float | datetime.datetime | None, ...]) -> dict[str, float]:
        """
        Parse a single row from the room sheet.

        These rows are often in the form "Length, X.XX, Width, Y.YY", so we iterate over those pairwise
        to extract the relevant fields (which can change by surface type, e.g. floors don't have a height).

        Parameters
        ----------
        row
            OpenPyxl row in the form of a tuple of row values

        Returns
        -------
        dict[str, float | None]
            Dictionary of element dimensions, can be None if not provided
        """
        result: dict[str, float] = {}
        for label, value in itertools.pairwise(row):
            if label == "Length (m)":
                length = try_convert_float(value)
                if length is not None:
                    result["length"] = length
            elif label == "Width (m)":
                width = try_convert_float(value)
                if width is not None:
                    result["width"] = width
            elif label == "Height (m)":
                height = try_convert_float(value)
                if height is not None:
                    result["height"] = height
            elif label == "Volume (m³)":
                volume = try_convert_float(value)
                volume = try_convert_float(value)
                if volume is not None:
                    result["volume"] = volume
            elif label == "U Value (W/m²K)":
                u_val = try_convert_float(value)
                if u_val is not None:
                    result["u_val"] = u_val

        if ("length" in result) and ("width" in result) and ("height" in result):
            result["volume"] = result["length"] * result["width"] * result["height"]
        elif ("length" in result) and ("width" in result):
            # It's a roof or floor
            result["area"] = result["length"] * result["width"]
        elif ("length" in result) and ("height" in result):
            result["area"] = result["length"] * result["height"]
        return result

    total_result = defaultdict(list)
    current_element = None
    for row in sheet.iter_rows(values_only=True, min_row=9, max_row=100):
        if "Room Dimensions" in row:
            current_element = "InternalAir"
            continue
        elif "Floor" in row:
            current_element = "Floor"
            continue
        elif "External Wall" in row:
            current_element = "ExternalWall"
            continue
        elif "Ceiling (Flat)" in row:
            current_element = "Ceiling"
            continue
        elif "Windows" in row:
            current_element = "Windows"
            continue
        elif "Roof Glazing" in row:
            current_element = "RoofGlazing"
            continue
        elif "External Doors" in row:
            current_element = "ExternalDoors"
            continue
        elif "Internal Walls" in row:
            current_element = "InternalWalls"
            continue

        if current_element is not None:
            parsed_row = parse_row(row)
            if not parsed_row.get("length") and not parsed_row.get("height") and not parsed_row.get("width"):
                continue
            total_result[current_element].append(parse_row(row))

    return dict(total_result)


def extract_rooms_metadata(sheet: openpyxl.worksheet.worksheet.Worksheet) -> MCSRoomMetadata:
    """
    Extract some metadata from the top box about a room.

    Parameters
    ----------
    sheet
        An MCS worksheet for a specific room

    Returns
    -------
    MCSRoomMetadata
        Name, level and air changes per hour for this room.
    """
    room_name = (sheet["C2"].value if sheet["C2"].value is not None else "") + (
        str(sheet["D2"].value) if sheet["D2"].value is not None else ""
    )
    if not room_name:
        room_name = None
    return MCSRoomMetadata(room_name=room_name, level=str(sheet["C3"].value), ach=try_convert_float(sheet["G2"].value))


def extract_rooms_data(fname: pathlib.Path) -> dict[str, float | dict[str, list[dict[str, float]]]]:
    """
    Extract the construction data including sizes and u values for all rooms in this survey.

    Parameters
    ----------
    fname
        An MCS survey spreadsheet

    Returns
    -------
    dict[room_name, dict[element, data]]
    """
    wb = openpyxl.load_workbook(fname, data_only=True, read_only=True, keep_links=False)

    all_results: dict[str, float | dict[str, list[dict[str, float]]]] = {}
    for sheet_name in wb.sheetnames:
        try:
            str(int(sheet_name))
        except ValueError:
            continue
        metadata = extract_rooms_metadata(wb[sheet_name])
        name = metadata["room_name"]

        room_result = parse_single_room(wb[sheet_name])
        if room_result and name:
            all_results[name] = room_result
            if metadata["ach"] is not None:
                all_results["air_changes"] = metadata["ach"]
    return all_results


def extract_u_values(fname: pathlib.Path) -> dict[str, float]:
    """
    Extract the U values table from an MCS document.

    This will include windows, floors and external doors of all types.
    The keys (material descriptions) are stripped of whitespace at either side, but
    otherwise retain their original (idisyncratic) capitalisation and descriptions.

    Parameters
    ----------
    fname
        MCS-type reporting spreadsheet

    Returns
    -------
        material_type: u value mapping for the materials in this survey.
    """
    wb = openpyxl.load_workbook(fname, data_only=True, read_only=True, keep_links=False)
    ws = wb["U Value Calculator"]
    external_material_idx, window_idx = None, None
    for idx, cell in enumerate(next(ws.rows)):
        if cell.value == "Wall/Roof/ceiling/non-ground floors U-values (SOURCE Domestic heating guide section 6)":
            external_material_idx = idx + 1
            break
    else:
        raise ValueError(f"Couldn't find Wall/Roof/ceiling properties in in {[item.value for item in next(ws.rows)]}")

    for idx, cell in enumerate(next(ws.rows)):
        if cell.value == "Window_description":
            window_idx = idx + 1
            break
    else:
        raise ValueError(f"Couldn't find Window_description in in {[item.value for item in next(ws.rows)]}")
    u_values_dict: dict[str, float] = {}

    # Check the external materials from their table
    for row in ws.iter_rows(min_col=external_material_idx, max_col=external_material_idx + 1, values_only=True, min_row=3):
        material = str(row[0]).strip() if row[0] is not None else None
        u_val = try_convert_float(row[1])
        if u_val is not None and material is not None:
            u_values_dict[material] = u_val
    # Check the external materials from their table
    for row in ws.iter_rows(min_col=window_idx, max_col=window_idx + 1, values_only=True, min_row=3):
        material = str(row[0]).strip() if row[0] is not None else None
        u_val = try_convert_float(row[1])
        if u_val is not None and material is not None:
            u_values_dict[material] = u_val
    return u_values_dict
