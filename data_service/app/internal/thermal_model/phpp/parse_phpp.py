"""Parse the PHPP spreadsheets to extract structural details."""

import datetime
import warnings
from collections.abc import Sequence
from decimal import Decimal
from pathlib import Path
from typing import BinaryIO, TypedDict

import numpy as np
import pandas as pd
from openpyxl.cell import Cell, MergedCell
from openpyxl.reader.excel import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.internal.thermal_model.heat_capacities import AIR_HEAT_CAPACITY
from app.internal.thermal_model.phpp.interventions import THIRD_PARTY_INTERVENTIONS, StructuralArea
from app.models.heating_load import FabricCostBreakdown

type ExcelRow = Sequence[Cell | MergedCell]


class StructuralRow(TypedDict):
    """A parsed PHPP row with structural data."""

    element_id: int
    name: str
    group: str
    area: float
    angle: float
    u_value: float
    area_type: StructuralArea


class StructuralInfo(TypedDict):
    """Metadata about a structure we got from a PHPP."""

    floor_area: float
    internal_volume: float
    air_changes: float


def maybe_cell_to_int(cell: Cell | MergedCell) -> int | None:
    """
    Try converting the contents of this cell to an integer.

    This will handle integer-like strings, and floats / decimals if they're exactly integers.

    Parameters
    ----------
    cell
        A cell with a value which is maybe an integer

    Returns
    -------
    int
        If the contents could be converted
    None
        If the contents are not integer-like
    """
    if isinstance(cell.value, str):
        try:
            # We often get indices like "001" which are technically strings
            # so let's turn them into an int
            return int(cell.value)
        except (ValueError, TypeError):
            # Couldn't convert this value to an int
            return None

    if isinstance(cell.value, int):
        return cell.value
    if isinstance(cell.value, float | Decimal):
        if int(cell.value) == cell.value:
            return int(cell.value)
        return None

    return None


def parse_phpp_area_row(row: ExcelRow) -> StructuralRow | None:
    """
    Parse a row from the PHPP areas sheet into a useful python dictionary.

    This will extract the information that we need from the row, such as U-values and areas.
    It is sensitive to the column indicies and is designed for PHPP v10.3.

    This will return None when we have met an expected failure.

    Parameters
    ----------
    row
        A row from the table with "Area input" at its top left corner

    Returns
    -------
    StructuralRow
        The details that we need from this row
    None
        If we failed to parse in an expected way
    """
    element_idx = maybe_cell_to_int(row[0])
    if element_idx is None:
        return None

    element_idx = element_idx
    element_name = str(row[1].value)
    element_group = str(row[2].value)

    total_size = row[15].value
    u_value = row[18].value
    if not isinstance(total_size, float) or not isinstance(u_value, float):
        return None

    angle = row[-1].value
    if not isinstance(angle, float):
        angle = float("NaN")

    if element_name.startswith("Wall"):
        area_type = StructuralArea.ExteriorWallArea
    elif element_name.startswith("Roof"):
        area_type = StructuralArea.RoofArea
    elif element_name.startswith("Floor"):
        area_type = StructuralArea.FloorArea
    elif element_name.startswith("Win"):
        raise ValueError("Tried to parse a window row with `parse_phpp_area_row`")
    else:
        raise ValueError(f"Got an unknown area type: {element_name}")
    return {
        "element_id": element_idx,
        "name": element_name,
        "group": element_group,
        "area": total_size,
        "angle": angle,
        "u_value": u_value,
        "area_type": area_type,
    }


def find_cell(
    ws: Worksheet,
    cell_contents: str,
    start_row: int = 0,
    end_row: int | None = None,
    start_col: int = 0,
    end_col: int | None = None,
) -> Cell:
    """
    Find a cell with a given contents in this worksheet.

    This searches row-by-row and column-by-column, returning the first entry we find in order.
    e.g. if there is that value in row 1, column 5 we will return that before we find the value in
    row 2, column 3. If you want to find all of these, re-run the search with start row > returned cell.row

    Parameters
    ----------
    ws
        Worksheet to scan thoruggh
    cell contents
        The string contents of the cell to check
    start_row
        Index of the first row to check (starts at 0)
    end_row
        Index of the last row to check (covers all rows by default)
    start_col
        Index of the first column to check (starts at 0)
    end_col
        Index of the last colimn to check (covers all columns by default)

    Returns
    -------
    Cell
        The cell we found, with a .row and .col_idx that you can use elsewhere.
    """
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col, values_only=False):
        for cell in row:
            if cell.value == cell_contents:
                # This check is actually fine, as MergedCells only have None cotnents
                assert isinstance(cell, Cell), "Must find value in a Cell not a MergedCell"
                return cell
    raise ValueError(f"Couldn't find cell contents {cell_contents} in {ws.title}")


def parse_phpp_window_row(row: ExcelRow) -> StructuralRow | None:
    """
    Parse a single row from the "windows" table.

    This will extract areas, associated walls and U-values for each window, with one window per
    row.

    Parameters
    ----------
    row
        A row from the windows table

    Returns
    -------
    StructuralRow
        A python dictionary with useful information extracted
    None
        If we couldn't parse this row correctly
    """
    element_idx = maybe_cell_to_int(row[0])
    if element_idx is None:
        return None
    element_name = str(row[1].value)
    if not element_name.startswith("Win"):
        raise ValueError("Tried to parse a non-window row with `parse_phpp_window_row`")

    element_group = str(row[7].value)

    total_size = row[36].value
    u_value = row[40].value
    if not isinstance(total_size, float) or not isinstance(u_value, float):
        return None

    angle = row[2].value
    if not isinstance(angle, float):
        angle = float("NaN")

    return {
        "element_id": element_idx,
        "name": element_name,
        "group": element_group,
        "angle": angle,
        "area": total_size,
        "u_value": u_value,
        "area_type": StructuralArea.WindowArea,
    }


def parse_phpp_thermal_bridge_row(row: ExcelRow) -> StructuralRow | None:
    """
    Parse a single row from the "windows" table.

    This will extract areas, associated walls and U-values for each window, with one window per
    row.

    Parameters
    ----------
    row
        A row from the windows table

    Returns
    -------
    StructuralRow
        A python dictionary with useful information extracted
    None
        If we couldn't parse this row correctly
    """
    element_idx = maybe_cell_to_int(row[0])
    if element_idx is None:
        return None
    # Note that thermal bridges might not have names
    element_name = str(row[1].value)
    if element_name == "-" or not element_name:
        element_name = "Default Thermal Bridge"

    element_group = str(row[2].value)

    total_length = row[9].value
    psi_value = row[11].value
    if not isinstance(total_length, float | int) or not isinstance(psi_value, float | int):
        return None

    return {
        "element_id": element_idx,
        "name": element_name,
        "group": element_group,
        "angle": float("NaN"),
        "area": float(total_length),
        "u_value": float(psi_value),
        "area_type": StructuralArea.ThermalBridge,
    }


def extract_phpp_ventilation(ws: Worksheet) -> tuple[float, float]:
    """
    Extract ventilation data including the total air volume and air change rate from this survey.

    Parameters
    ----------
    ws
        The worksheet, probably titled "Ventilation"

    Returns
    -------
    tuple[float, float]
        Air volume in m^3m air changes per hour
    """
    vent_left_cell = find_cell(ws, "Reference volume for the ventilation system (ATFA*h) =")
    volume_cell = ws[vent_left_cell.row][vent_left_cell.col_idx + 3]
    if isinstance(volume_cell.value, str | float | int):
        air_volume = float(volume_cell.value)
    else:
        raise TypeError(f"Got bad value for volume {volume_cell.value}")

    ach_left_cell = find_cell(ws, "Air change rate from pressurisation test")
    ach_cell = ws[ach_left_cell.row][ach_left_cell.col_idx + 3]
    if isinstance(ach_cell.value, str | float | int):
        ach = float(ach_cell.value)
    else:
        raise TypeError(f"Got bad value for ACH {ach_cell.value}")
    return air_volume, ach


def apply_phpp_intervention(structure_df: pd.DataFrame, intervention_name: str) -> pd.DataFrame:
    """
    Apply an intervention to a PHPP structure.

    Interventions act on areas, which have been tagged in the structural dataframe.
    We'll look up the new U-value, and apply the intervention to all elements of that type with a U-value
    worse than the new one.

    Parameters
    ----------
    structure_df
        Structural df with `area_type` and `u_value` columns
    intervention_name
        Name of the intervention to match the list in interventions.py

    Returns
    -------
    pd.DataFrame
        Structural dataframe with intervention applied
    """
    # Don't clobber the old structure!
    new_df = structure_df.copy()

    if intervention_name not in THIRD_PARTY_INTERVENTIONS:
        raise ValueError(f"Bad intervention `{intervention_name}`; check THIRD_PARTY_INTERVENTIONS for a list of good ones.")
    intervention_u_value = THIRD_PARTY_INTERVENTIONS[intervention_name]["u_value"]
    is_right_area = new_df["area_type"].isin(THIRD_PARTY_INTERVENTIONS[intervention_name]["acts_on"])
    is_worse_u_value = new_df["u_value"] >= intervention_u_value

    both_mask = np.logical_and(is_right_area, is_worse_u_value)
    new_df.loc[both_mask, "u_value"] = intervention_u_value

    return new_df


def phpp_fabric_intervention_cost(
    structure_df: pd.DataFrame, interventions: list[str]
) -> tuple[float, list[FabricCostBreakdown]]:
    """
    Calculate the cost of a combined set of interventions.

    Interventions act on areas, which have been tagged in the structural dataframe.
    We'll look up the new U-value, and apply the intervention to all elements of that type with a U-value
    worse than the new one. Then, we'll calculate the cost of all the best interventions for each element,
    assuming no redundant work.

    Parameters
    ----------
    structure_df
        Structural df with `area_type` and `u_value` columns
    intervention_name
        Name of the intervention to match the list in interventions.py

    Returns
    -------
    float
        Total cost of all interventionns
    """
    # Don't clobber the old structure, but create a new one tracking costs which are zeroed out for unaffected elements.
    new_df = structure_df.copy()
    new_df["cost"] = 0.0
    new_df["intervention"] = None

    # First, assign all the "structural" interventions that replace fabric interventions.
    for intervention_name in interventions:
        if intervention_name not in THIRD_PARTY_INTERVENTIONS:
            raise ValueError(f"Bad intervention `{intervention_name}`; check THIRD_PARTY_INTERVENTIONS for a list of good ones.")
        intervention_u_value = THIRD_PARTY_INTERVENTIONS[intervention_name]["u_value"]
        if intervention_u_value is None:
            continue
        intervention_cost = THIRD_PARTY_INTERVENTIONS[intervention_name]["cost"]
        is_right_area = new_df["area_type"].isin(THIRD_PARTY_INTERVENTIONS[intervention_name]["acts_on"])
        is_worse_u_value = new_df["u_value"] >= intervention_u_value
        both_mask = np.logical_and(is_right_area, is_worse_u_value)
        new_df.loc[both_mask, "cost"] = intervention_cost
        new_df.loc[both_mask, "intervention"] = intervention_name

    # We calculate the breakdown after we've applied all the interventions as they might overwrite each other.
    # However, we also apply the "non-structural" interventions here.
    fabric_cost_breakdown: list[FabricCostBreakdown] = []
    for intervention_name in interventions:
        # Where we had a None U-value, e.g. for air tightness, there's still an affected area. This can't be overwritten
        # by structural interventions, so take the total area of that type to base our costing off.
        if THIRD_PARTY_INTERVENTIONS[intervention_name]["u_value"] is None:
            mask = new_df["area_type"].isin(THIRD_PARTY_INTERVENTIONS[intervention_name]["acts_on"])
        else:
            mask = new_df["intervention"] == intervention_name

        affected_df = new_df[mask]
        fabric_cost_breakdown.append(
            FabricCostBreakdown(
                name=intervention_name,
                cost=(THIRD_PARTY_INTERVENTIONS[intervention_name]["cost"] * affected_df["area"]).sum(),
                area=affected_df["area"].sum(),
            )
        )

    return sum(item.cost for item in fabric_cost_breakdown), fabric_cost_breakdown


def phpp_to_dataframe(fpath: Path | BinaryIO) -> tuple[pd.DataFrame, StructuralInfo]:
    """
    Turn a PHPP stored in an Excel file into a pandas dataframe with useful information extracted.

    The PHPP format is expected to be v10.3 and the process is relatively fragile: we use hard coded indices to get
    the right columns from each sheet.
    Currently just extracts windows, roof, floor, and walls.

    Parameters
    ----------
    fpath
        Path to the PHPP excel file, or file-like object with it already opened in binary mode

    Returns
    -------
    pd.DataFrame
        Pandas dataframe with "element_id", "name", "group", "angle", "area" and "u_value" columns.
    StructuralInfo
        other stuff about a structure that isn't necessarily a fabric element, like volume and ACH
    """
    with warnings.catch_warnings(category=UserWarning, action="ignore"):
        wb = load_workbook(fpath, data_only=True, keep_links=False)

    ws = wb["Areas"]

    area_header = find_cell(ws, "Area input")
    all_rows: list[StructuralRow] = []
    for row in ws.iter_rows(min_row=area_header.row, min_col=area_header.col_idx):
        if row[area_header.col_idx].value == "<End of designPH import!>":
            break

        try:
            parsed_row = parse_phpp_area_row(row)
        except ValueError:
            # We got a bad entry
            parsed_row = None
        if parsed_row is not None:
            all_rows.append(parsed_row)

    window_sheet = wb["Windows"]
    window_start = find_cell(window_sheet, "Description")
    for row in window_sheet.iter_rows(min_row=window_start.row + 2, min_col=window_start.col_idx):
        try:
            parsed_row = parse_phpp_window_row(row)
        except ValueError:
            # This row isn't parseable, so we've run out of good windows
            # This is an expected outcome, it's cheaper to throw an exception once we fail to get a single cell
            break
        if parsed_row is not None:
            all_rows.append(parsed_row)

    thermal_bridge_start = find_cell(wb["Areas"], "Thermal bridge input")
    for row in ws.iter_rows(min_row=thermal_bridge_start.row, min_col=thermal_bridge_start.col_idx):
        if row[0].value == "<End of designPH import!>":
            break

        try:
            parsed_row = parse_phpp_thermal_bridge_row(row)
        except ValueError:
            parsed_row = None

        if parsed_row is not None:
            all_rows.append(parsed_row)
    df = pd.DataFrame.from_records(all_rows).set_index("element_id")
    internal_volume, ach = extract_phpp_ventilation(wb["Ventilation"])
    return (df, StructuralInfo(floor_area=1.0, internal_volume=internal_volume, air_changes=ach))


def phpp_fabric_heat_loss(structure_df: pd.DataFrame, internal_t: float = 21, external_t: float = -2.3) -> float:
    """
    Calculate the fabric heat loss from this PHPP survey.

    The fabric heat loss is the amount of heat that flows through fabric elements excluding air changes.
    It is proportional to the temperature difference between inside and out, with coefficient of proportionality
    being U value times area for each element.

    Parameters
    ----------
    structure_df
        Structural dataframe including walls, windows, floors etc with specified aras and u values
    internal_t
        Internal thermostat temperature in °C, probably 21ish
    external_t
        External air temperature in °C

    Returns
    -------
    float
        Heat loss in Watts
    """
    total_heat_loss = 0.0
    delta_t = internal_t - external_t
    for area, u_value in zip(structure_df["area"], structure_df["u_value"], strict=True):
        total_heat_loss += area * u_value * delta_t

    return total_heat_loss


def phpp_total_heat_loss(
    structure_df: pd.DataFrame, metadata: StructuralInfo, internal_t: float = 21, external_t: float = -2.3
) -> float:
    """
    Calculate the total heat loss including fabric and convection components.

    This combines the fabric heat loss with an extra component representing air changes. It does not include solar gains,
    as for peak heat loss calculations we presume that there are no gains.

    Parameters
    ----------
    structure_df
        Structural dataframe including walls, windows, floors etc with specified aras and u values
    metadata
        Metadata about the building including its air volume
    internal_t
        Internal thermostat temperature in °C, probably 21ish
    external_t
        External air temperature in °C

    Returns
    -------
    float
        Heat loss in Watts
    """
    fabric_component = phpp_fabric_heat_loss(structure_df, internal_t, external_t)
    # Air changes come in air changes per hour, so convert this into Watts
    convection_component = (
        metadata["air_changes"] * metadata["internal_volume"] * AIR_HEAT_CAPACITY / datetime.timedelta(hours=1).total_seconds()
    )
    return fabric_component + convection_component
